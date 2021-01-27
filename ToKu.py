import json
import pathlib
import re
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

confirm_query = "^[Tt1Yy]|[Tt][Aa][Kk]$"
reg_file_name_json = "^.+\\.json$"
reg_file_name_xlsx = "^.+\\.xlsx$"

def check_if_file_exist(file_name, quiet=False):
    file = pathlib.Path(file_name)
    if file.exists():
        if quiet:
            return True

        confirm = input("Plik isnieje. Chcesz go nadpisać? (t - tak):")
        if not re.search(confirm_query, confirm):
            return True
    return False

def add_sheet_to_workbook(table: List[dict], wb=Workbook(), title="Untitled", overwrite=False):
    if overwrite:
        ws = wb.worksheets[-1]
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    for cell in table:
        ws.cell(row=cell['row'], column=cell['column']).value = cell['value']
    return ws

def save_json_to_xlsx():
    file_name = input("Wpisz nazwe pliku z rozszerzeniem json: ")
    if not re.search(reg_file_name_json, file_name):
        file_name = "{}.json".format(file_name)

    try:
        with open("{}".format(file_name), "r") as json_file:
            tables = json.load(json_file)
            wb = Workbook()

            file_name_out = input("Wpisz nazwę pliku, do którego zapisać dane: ")
            if not re.search(reg_file_name_xlsx, file_name_out):
                file_name_out="{}.xslx".format(file_name_out)

            overwrite_first_sheet = False
            if check_if_file_exist(file_name_out, True):
                wb = load_workbook(file_name_out)
            else:
                overwrite_first_sheet = True

            if type(tables) == list:
                add_sheet_to_workbook(tables, wb, overwrite=overwrite_first_sheet)
            elif type(tables) == dict:
                for table_keys in tables.keys():
                    add_sheet_to_workbook(tables[table_keys], wb, table_keys, overwrite_first_sheet)
                    overwrite_first_sheet = False
            del overwrite_first_sheet

            wb.save(file_name_out)
    except InvalidFileException:
        print("Nieobsługiwany format pliku!")
    except PermissionError as e:
        print("Brak uprawnień do pliku! ", e.filename)
    except FileNotFoundError as e:
        print("Nie znaleziono pliku! ", e.filename)

save_json_to_xlsx()